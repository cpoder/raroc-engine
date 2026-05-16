"""Build the three multi-period golden fixtures for Task 1.1.

Independent implementation of the per-period RAROC math (no import
from raroc_engine.calculator — that's the system under test). Uses
scipy.stats.norm and python math only.

Outputs:
- tests/fixtures/period_rcf_5y.yaml
- tests/fixtures/period_termloan_7y_amortising.yaml
- tests/fixtures/period_projfin_10y_grace.yaml
- tests/fixtures/reference_excel/period_rcf_5y.xlsx
- tests/fixtures/reference_excel/period_termloan_7y_amortising.xlsx
- tests/fixtures/reference_excel/period_projfin_10y_grace.xlsx

Run:
    python tests/fixtures/build_fixtures.py

Re-running is idempotent. The YAML is the test contract; the Excel
is the human-readable reference (formulas + computed values).
"""

from __future__ import annotations

import math
import os
from dataclasses import dataclass, field, asdict
from datetime import date, timedelta
from typing import List, Dict, Optional

import yaml
from scipy.stats import norm

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ──────────────────────────────────────────────────────────────────
# Reference data (mirrors repository/ratings.csv + engine defaults)
# ──────────────────────────────────────────────────────────────────

PD_BY_RATING: Dict[str, float] = {
    "Aaa": 0.0001, "Aa1": 0.0001, "Aa2": 0.0001, "Aa3": 0.0003,
    "A1":  0.0004, "A2":  0.0005, "A3":  0.0007,
    "Baa1": 0.0010, "Baa2": 0.0016, "Baa3": 0.0024,
    "Ba1":  0.0038, "Ba2":  0.0063, "Ba3":  0.0111,
    "B1":   0.0214, "B2":   0.0382, "B3":   0.0712,
    "Caa1": 0.1500, "Caa2": 0.2682, "Caa3": 0.3500,
    "Ca":   0.5000, "C":    0.9999,
}


@dataclass
class EngineConfigSnapshot:
    """Mirror of raroc_engine.config.EngineConfig defaults used in fixtures."""
    regime: str = "basel3"
    risk_free_rate: float = 0.0325
    bank_tax_rate: float = 0.25
    funding_cost_bp: float = 0.0
    default_cost_income_ratio: float = 0.40
    output_floor_pct: float = 0.55
    pd_floor: float = 0.0005
    lgd_floor_unsecured: float = 0.25
    lgd_floor_secured: float = 0.10
    default_collateral_type: str = "receivables"


@dataclass
class Period:
    index: int
    start: date
    end: date
    dt_years: float
    commitment: float
    avg_drawn: float
    remaining_maturity_years: float
    upfront_fee: float = 0.0
    flat_fee: float = 0.0
    participation_fee: float = 0.0
    floating_index: Optional[str] = None
    fixing_rate: Optional[float] = None


@dataclass
class DealStatic:
    fixture_id: str
    description: str
    product_type: str
    currency: str
    rating: str
    global_grr: float
    confirmed: bool
    spread: float
    commitment_fee: float


@dataclass
class DiscountSpec:
    kind: str = "scalar"
    rate: float = 0.0325
    day_count: str = "Act/365F"


# ──────────────────────────────────────────────────────────────────
# Math (independent of raroc_engine — system under test)
# ──────────────────────────────────────────────────────────────────

def sa_risk_weight(pd_val: float) -> float:
    """Basel III SA risk weight, corporates (BIS d424 Table 5)."""
    if pd_val <= 0.0005:
        return 0.20
    if pd_val <= 0.0015:
        return 0.50
    if pd_val <= 0.0075:
        return 0.75
    if pd_val <= 0.03:
        return 1.00
    return 1.50


def lgd_floor_for(grr: float, cfg: EngineConfigSnapshot) -> float:
    """Match calculator.py: coll_type='none' if grr==0 else default ('receivables')."""
    if grr == 0:
        return cfg.lgd_floor_unsecured
    if cfg.default_collateral_type in ("receivables", "real_estate"):
        return cfg.lgd_floor_secured
    return cfg.lgd_floor_unsecured


def confirmed_exposure(commitment: float, drawn: float) -> float:
    """CCF: confirmed MLT = 0.25 * drawn + 0.75 * commitment."""
    return 0.25 * drawn + 0.75 * commitment


def discount_factor(t_years: float, rate: float) -> float:
    return (1.0 + rate) ** (-t_years)


@dataclass
class PeriodRow:
    """One row of the period engine output."""
    index: int
    start: date
    end: date
    dt_years: float
    commitment: float
    avg_drawn: float
    remaining_maturity_years: float

    # Revenue / cost
    revenue: float = 0.0
    cost: float = 0.0
    funding_cost: float = 0.0

    # Risk
    exposure: float = 0.0
    pd: float = 0.0
    pd_basel2: float = 0.0
    lgd: float = 0.0
    correlation: float = 0.0  # R
    maturity_adj_b: float = 0.0
    z: float = 0.0
    K_irb: float = 0.0
    sa_rw: float = 0.0
    K_floor: float = 0.0
    K: float = 0.0
    fpe: float = 0.0
    el: float = 0.0

    # Margins
    gross_margin: float = 0.0
    fpe_return: float = 0.0
    net_margin: float = 0.0
    raroc: float = 0.0

    # Discount
    t_end_years: float = 0.0
    df: float = 0.0
    revenue_pv: float = 0.0
    net_margin_pv: float = 0.0
    drawn_pv: float = 0.0


def compute_period(
    p: Period,
    deal: DealStatic,
    cfg: EngineConfigSnapshot,
    cumulative_t_end: float,
    disc_rate: float,
) -> PeriodRow:
    """Independent implementation of one period of multi-period RAROC."""

    row = PeriodRow(
        index=p.index,
        start=p.start,
        end=p.end,
        dt_years=p.dt_years,
        commitment=p.commitment,
        avg_drawn=p.avg_drawn,
        remaining_maturity_years=p.remaining_maturity_years,
    )

    # 1. Revenue (period-scaled spread + commit, plus period-allocated fees)
    row.revenue = (
        deal.spread * p.avg_drawn * p.dt_years
        + deal.commitment_fee * (p.commitment - p.avg_drawn) * p.dt_years
        + p.upfront_fee + p.flat_fee + p.participation_fee
    )

    # 2. Cost (cost-of-income ratio)
    row.cost = row.revenue * cfg.default_cost_income_ratio

    # 3. Exposure (confirmed MLT CCF)
    row.exposure = confirmed_exposure(p.commitment, p.avg_drawn)

    # 4. PD (rating, floored under Basel III)
    pd_raw = PD_BY_RATING[deal.rating]
    if cfg.regime == "basel3":
        row.pd = max(pd_raw, cfg.pd_floor)
    else:
        row.pd = pd_raw
    row.pd_basel2 = row.pd * (1.0 - deal.global_grr)

    # 5. LGD (with floor)
    lgd_raw = 1.0 - deal.global_grr
    floor = lgd_floor_for(deal.global_grr, cfg)
    row.lgd = max(lgd_raw, floor)

    # 6. Asset correlation R (BIS CRE31)
    row.correlation = (
        0.12 * (1.0 + math.exp(-50.0 * row.pd) - 2.0 * math.exp(-50.0))
        / (1.0 - math.exp(-50.0))
    )

    # 7. Maturity adjustment b
    row.maturity_adj_b = (0.11852 - 0.05478 * math.log(row.pd)) ** 2

    # 8. Capital K (IRB) with output floor
    M = p.remaining_maturity_years
    row.z = (
        math.sqrt(1.0 / (1.0 - row.correlation)) * norm.ppf(row.pd)
        + math.sqrt(row.correlation / (1.0 - row.correlation)) * norm.ppf(0.999)
    )
    row.K_irb = (
        row.lgd * (norm.cdf(row.z) - row.pd)
        * (1.0 + (M - 2.5) * row.maturity_adj_b)
        / (1.0 - 1.5 * row.maturity_adj_b)
    )
    if cfg.regime == "basel3":
        row.sa_rw = sa_risk_weight(row.pd)
        row.K_floor = cfg.output_floor_pct * row.sa_rw / 12.5
        row.K = max(row.K_irb, row.K_floor)
    else:
        row.sa_rw = 0.0
        row.K_floor = 0.0
        row.K = row.K_irb

    # 9. FPE, EL
    row.fpe = row.exposure * row.K
    row.el = row.exposure * row.pd_basel2 * p.dt_years
    row.funding_cost = cfg.funding_cost_bp * row.exposure * p.dt_years

    # 10. Margins
    row.gross_margin = row.revenue - row.cost - row.funding_cost
    row.fpe_return = cfg.risk_free_rate * row.fpe * p.dt_years
    row.net_margin = row.gross_margin - row.el + row.fpe_return

    # 11. RAROC (matches calculator step 11)
    if row.fpe > 0:
        row.raroc = (1.0 - cfg.bank_tax_rate) * (
            (row.revenue - row.cost - row.funding_cost - row.el) / row.fpe
            + cfg.risk_free_rate
        )

    # 12. Discount (end-of-period convention)
    row.t_end_years = cumulative_t_end + p.dt_years
    row.df = discount_factor(row.t_end_years, disc_rate)
    row.revenue_pv = row.revenue * row.df
    row.net_margin_pv = row.net_margin * row.df
    row.drawn_pv = p.avg_drawn * p.dt_years * row.df

    return row


def compute_aggregates(rows: List[PeriodRow]) -> Dict[str, float]:
    npv_borrower_cost = sum(r.revenue_pv for r in rows)
    npv_bank_net_margin = sum(r.net_margin_pv for r in rows)
    npv_drawn_balance = sum(r.drawn_pv for r in rows)
    fpe_dt_sum = sum(r.fpe * r.dt_years for r in rows)
    fpe_weighted_raroc = (
        sum(r.raroc * r.fpe * r.dt_years for r in rows) / fpe_dt_sum
        if fpe_dt_sum > 0 else 0.0
    )
    effective_spread = (
        npv_borrower_cost / npv_drawn_balance if npv_drawn_balance > 0 else 0.0
    )
    total_revenue = sum(r.revenue for r in rows)
    total_el = sum(r.el for r in rows)
    avg_exposure = (
        sum(r.exposure * r.dt_years for r in rows) / sum(r.dt_years for r in rows)
    )
    return {
        "npv_borrower_cost": npv_borrower_cost,
        "npv_bank_net_margin": npv_bank_net_margin,
        "npv_drawn_balance": npv_drawn_balance,
        "effective_spread": effective_spread,
        "effective_spread_bp": effective_spread * 10000.0,
        "fpe_weighted_raroc": fpe_weighted_raroc,
        "total_revenue_undisc": total_revenue,
        "total_el_undisc": total_el,
        "avg_exposure": avg_exposure,
    }


# ──────────────────────────────────────────────────────────────────
# Fixture definitions
# ──────────────────────────────────────────────────────────────────

START_DATE = date(2026, 6, 1)


def annual_periods(
    commitments: List[float],
    avg_drawns: List[float],
    *,
    start: date = START_DATE,
    upfront_fee_period1: float = 0.0,
) -> List[Period]:
    """Build a list of annual Period rows from parallel lists.

    Uses a stylized dt_years=1.0 for every period (matches the existing
    single-period calculator's 1-year assumption and removes leap-year
    noise from the reference fixtures). Period engine tests for Act/365F
    day-count behaviour come in a separate sub-annual fixture in Q2.
    """
    n = len(commitments)
    assert len(avg_drawns) == n
    periods = []
    cursor = start
    for i in range(n):
        nxt = date(cursor.year + 1, cursor.month, cursor.day)
        periods.append(
            Period(
                index=i + 1,
                start=cursor,
                end=nxt,
                dt_years=1.0,
                commitment=commitments[i],
                avg_drawn=avg_drawns[i],
                remaining_maturity_years=float(n - i),
                upfront_fee=upfront_fee_period1 if i == 0 else 0.0,
            )
        )
        cursor = nxt
    return periods


def fixture_rcf_5y():
    deal = DealStatic(
        fixture_id="period_rcf_5y",
        description=(
            "Confirmed 5y RCF, 50M commitment, cleandown to 20M drawn after year 3. "
            "Investment-grade mid-cap borrower (Baa2), unsecured. 150bp spread, "
            "25bp commit fee on undrawn, 200k upfront fee in year 1."
        ),
        product_type="mlt_credit",
        currency="EUR",
        rating="Baa2",
        global_grr=0.0,
        confirmed=True,
        spread=0.0150,
        commitment_fee=0.0025,
    )
    periods = annual_periods(
        commitments=[50_000_000] * 5,
        avg_drawns=[35_000_000, 35_000_000, 35_000_000, 20_000_000, 20_000_000],
        upfront_fee_period1=200_000,
    )
    return deal, periods


def fixture_termloan_7y_amortising():
    deal = DealStatic(
        fixture_id="period_termloan_7y_amortising",
        description=(
            "7y bullet-drawn amortising term loan, 70M day-1 drawn, 10M/yr "
            "linear paydown to 0. Crossover-credit borrower (Baa3), partially "
            "secured (GRR 0.20). 175bp spread, no commit fee (fully drawn day 1), "
            "350k upfront fee in year 1."
        ),
        product_type="mlt_credit",
        currency="EUR",
        rating="Baa3",
        global_grr=0.20,
        confirmed=True,
        spread=0.0175,
        commitment_fee=0.0,
    )
    # Linear amortisation: 70 → 60 → 50 → 40 → 30 → 20 → 10 → 0
    # avg_drawn per year = (start + end) / 2
    starts = [70, 60, 50, 40, 30, 20, 10]
    ends = [60, 50, 40, 30, 20, 10, 0]
    avg = [(s + e) / 2 * 1_000_000 for s, e in zip(starts, ends)]
    commitments = [s * 1_000_000 for s in starts]
    periods = annual_periods(
        commitments=commitments,
        avg_drawns=avg,
        upfront_fee_period1=350_000,
    )
    return deal, periods


def fixture_projfin_10y_grace():
    deal = DealStatic(
        fixture_id="period_projfin_10y_grace",
        description=(
            "10y project finance: 100M commitment, 3y drawdown ramp (30→70→100M), "
            "2y grace (100M drawn), 4y amortise (100→20M), bullet 20M repaid "
            "end-y10. BB+ sponsor (Ba1), project-asset security (GRR 0.30). "
            "225bp spread, 35bp commit fee on undrawn portion, "
            "1M upfront in year 1."
        ),
        product_type="mlt_credit",
        currency="EUR",
        rating="Ba1",
        global_grr=0.30,
        confirmed=True,
        spread=0.0225,
        commitment_fee=0.0035,
    )
    # Drawdown profile (avg drawn over year)
    avg_drawn = [
        30_000_000,  # y1 ramp
        70_000_000,  # y2 ramp
        100_000_000, # y3 fully drawn
        100_000_000, # y4 grace
        100_000_000, # y5 grace
        90_000_000,  # y6 amort (100→80)
        70_000_000,  # y7 amort (80→60)
        50_000_000,  # y8 amort (60→40)
        30_000_000,  # y9 amort (40→20)
        20_000_000,  # y10 bullet (20→0 at end)
    ]
    commitments = [100_000_000] * 10
    periods = annual_periods(
        commitments=commitments,
        avg_drawns=avg_drawn,
        upfront_fee_period1=1_000_000,
    )
    return deal, periods


# ──────────────────────────────────────────────────────────────────
# YAML writer
# ──────────────────────────────────────────────────────────────────

def _r(x: float, places: int = 8) -> float:
    """Round to a fixed precision for YAML readability while keeping tolerances comfortable.

    scipy.stats returns numpy scalars; coerce to plain float so PyYAML's
    safe representer accepts them.
    """
    return float(round(float(x), places))


def write_yaml(
    deal: DealStatic,
    periods: List[Period],
    rows: List[PeriodRow],
    aggregates: Dict[str, float],
    cfg: EngineConfigSnapshot,
    discount: DiscountSpec,
    out_path: str,
) -> None:
    doc = {
        "fixture_id": deal.fixture_id,
        "description": deal.description,
        "version": "0.1",
        "generated_by": "tests/fixtures/build_fixtures.py",
        "engine_config": {
            "regime": cfg.regime,
            "risk_free_rate": cfg.risk_free_rate,
            "bank_tax_rate": cfg.bank_tax_rate,
            "funding_cost_bp": cfg.funding_cost_bp,
            "default_cost_income_ratio": cfg.default_cost_income_ratio,
            "output_floor_pct": cfg.output_floor_pct,
            "pd_floor": cfg.pd_floor,
            "lgd_floor_unsecured": cfg.lgd_floor_unsecured,
            "lgd_floor_secured": cfg.lgd_floor_secured,
            "default_collateral_type": cfg.default_collateral_type,
        },
        "deal": {
            "product_type": deal.product_type,
            "currency": deal.currency,
            "rating": deal.rating,
            "global_grr": deal.global_grr,
            "confirmed": deal.confirmed,
            "spread": deal.spread,
            "commitment_fee": deal.commitment_fee,
        },
        "discount": {
            "kind": discount.kind,
            "rate": discount.rate,
            "day_count": discount.day_count,
        },
        "schedule": {
            "type": "annual",
            "day_count": "Act/365F",
            "periods": [
                {
                    "index": p.index,
                    "start": p.start.isoformat(),
                    "end": p.end.isoformat(),
                    "dt_years": _r(p.dt_years, 10),
                    "commitment": _r(p.commitment, 2),
                    "avg_drawn": _r(p.avg_drawn, 2),
                    "remaining_maturity_years": _r(p.remaining_maturity_years, 4),
                    "upfront_fee": _r(p.upfront_fee, 2),
                    "flat_fee": _r(p.flat_fee, 2),
                    "participation_fee": _r(p.participation_fee, 2),
                    "floating_index": p.floating_index,
                    "fixing_rate": p.fixing_rate,
                }
                for p in periods
            ],
        },
        "expected": {
            "per_period": [
                {
                    "index": r.index,
                    "revenue": _r(r.revenue),
                    "cost": _r(r.cost),
                    "funding_cost": _r(r.funding_cost),
                    "exposure": _r(r.exposure),
                    "pd": _r(r.pd, 10),
                    "pd_basel2": _r(r.pd_basel2, 10),
                    "lgd": _r(r.lgd, 8),
                    "correlation_R": _r(r.correlation, 10),
                    "maturity_adj_b": _r(r.maturity_adj_b, 10),
                    "z": _r(r.z, 10),
                    "K_irb": _r(r.K_irb, 10),
                    "sa_rw": _r(r.sa_rw, 6),
                    "K_floor": _r(r.K_floor, 10),
                    "K": _r(r.K, 10),
                    "fpe": _r(r.fpe),
                    "el": _r(r.el),
                    "gross_margin": _r(r.gross_margin),
                    "fpe_return": _r(r.fpe_return),
                    "net_margin": _r(r.net_margin),
                    "raroc": _r(r.raroc, 10),
                    "raroc_bp": _r(r.raroc * 10000.0, 4),
                    "t_end_years": _r(r.t_end_years, 10),
                    "df": _r(r.df, 10),
                    "revenue_pv": _r(r.revenue_pv),
                    "net_margin_pv": _r(r.net_margin_pv),
                    "drawn_pv": _r(r.drawn_pv),
                }
                for r in rows
            ],
            "aggregates": {k: _r(v, 8) for k, v in aggregates.items()},
            "discount_meta": {
                "curve_status": "scalar",
                "rate_used": discount.rate,
                "day_count": discount.day_count,
            },
        },
        "tolerances": {
            "per_period_raroc_bp_abs": 0.5,
            "per_period_fpe_rel": 0.005,
            "npv_rel": 0.001,
            "effective_spread_bp_abs": 0.5,
            "single_period_parity_abs": 1e-12,
        },
    }

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w") as f:
        yaml.safe_dump(doc, f, default_flow_style=False, sort_keys=False, width=120)


# ──────────────────────────────────────────────────────────────────
# Excel writer
# ──────────────────────────────────────────────────────────────────

def _bold(cell, color: str = "FFFFFF", bg: str = "1F4E78"):
    cell.font = Font(bold=True, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)


def write_excel(
    deal: DealStatic,
    periods: List[Period],
    rows: List[PeriodRow],
    aggregates: Dict[str, float],
    cfg: EngineConfigSnapshot,
    discount: DiscountSpec,
    out_path: str,
) -> None:
    if not HAS_OPENPYXL:
        print(f"[warn] openpyxl missing, skipping {out_path}")
        return

    wb = Workbook()

    # ── Sheet 1: Inputs ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Inputs"

    rows_in = [
        ("Fixture ID", deal.fixture_id),
        ("Description", deal.description),
        ("Currency", deal.currency),
        ("Product type", deal.product_type),
        ("", ""),
        ("─ Engine config ─", ""),
        ("Regime", cfg.regime),
        ("Risk-free rate", cfg.risk_free_rate),
        ("Bank tax rate", cfg.bank_tax_rate),
        ("Funding cost (decimal)", cfg.funding_cost_bp),
        ("Cost-income ratio", cfg.default_cost_income_ratio),
        ("Output floor %", cfg.output_floor_pct),
        ("PD floor", cfg.pd_floor),
        ("LGD floor — unsecured", cfg.lgd_floor_unsecured),
        ("LGD floor — secured", cfg.lgd_floor_secured),
        ("Default collateral type", cfg.default_collateral_type),
        ("", ""),
        ("─ Deal static ─", ""),
        ("Rating", deal.rating),
        (f"PD ({deal.rating}) raw", PD_BY_RATING[deal.rating]),
        ("Global GRR", deal.global_grr),
        ("Confirmed", deal.confirmed),
        ("Spread", deal.spread),
        ("Commitment fee", deal.commitment_fee),
        ("", ""),
        ("─ Discount ─", ""),
        ("Kind", discount.kind),
        ("Rate", discount.rate),
        ("Day count", discount.day_count),
        ("", ""),
        ("─ Formula notes ─", ""),
        ("PD (after floor)", "=MAX(PD_raw, PD_floor)"),
        ("LGD", "=MAX(1-GRR, LGD_floor)  (floor depends on collateral type)"),
        ("Exposure (confirmed MLT)", "=0.25*avg_drawn + 0.75*commitment"),
        ("R", "=0.12 * (1 + EXP(-50*PD) - 2*EXP(-50)) / (1 - EXP(-50))"),
        ("b", "=(0.11852 - 0.05478*LN(PD))^2"),
        ("z", "=SQRT(1/(1-R))*NORM.S.INV(PD) + SQRT(R/(1-R))*NORM.S.INV(0.999)"),
        ("K_irb", "=LGD * (NORM.S.DIST(z, TRUE) - PD) * (1 + (M-2.5)*b) / (1 - 1.5*b)"),
        ("SA_RW", "stepwise: 0.20|0.50|0.75|1.00|1.50 by PD bucket"),
        ("K_floor", "=output_floor_pct * SA_RW / 12.5"),
        ("K", "=MAX(K_irb, K_floor)"),
        ("FPE", "=exposure * K"),
        ("EL", "=exposure * PD * (1-GRR) * dt_years"),
        ("Revenue", "=spread*avg_drawn*dt + commit_fee*(commitment-avg_drawn)*dt + period_fees"),
        ("Cost", "=revenue * cost_income_ratio"),
        ("Gross margin", "=revenue - cost - funding_cost"),
        ("FPE return", "=risk_free_rate * FPE * dt_years"),
        ("Net margin", "=gross_margin - EL + fpe_return"),
        ("RAROC", "=(1-tax) * ((revenue-cost-funding_cost-EL)/FPE + risk_free_rate)"),
        ("DF (end of period)", "=(1 + disc_rate) ^ (-t_end_years)"),
    ]

    for i, (k, v) in enumerate(rows_in, start=1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        if k.startswith("─"):
            _bold(ws.cell(row=i, column=1))

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100

    # ── Sheet 2: Periods ────────────────────────────────────────
    ws2 = wb.create_sheet("Periods")
    headers = [
        "Period", "Start", "End", "dt_yr", "Commitment", "AvgDrawn",
        "M_remain_yr", "Upfront", "Revenue", "Cost", "Exposure",
        "PD", "PD_basel2", "LGD", "R", "b", "z", "K_irb",
        "SA_RW", "K_floor", "K", "FPE", "EL", "FundingCost",
        "GrossMargin", "FPEReturn", "NetMargin", "RAROC", "RAROC_bp",
        "t_end_yr", "DF", "Revenue_PV", "NetMargin_PV", "Drawn_PV",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        _bold(c)

    for ridx, r in enumerate(rows, start=2):
        vals = [
            r.index, r.start.isoformat(), r.end.isoformat(), r.dt_years,
            r.commitment, r.avg_drawn, r.remaining_maturity_years,
            periods[r.index - 1].upfront_fee,
            r.revenue, r.cost, r.exposure,
            r.pd, r.pd_basel2, r.lgd, r.correlation, r.maturity_adj_b, r.z, r.K_irb,
            r.sa_rw, r.K_floor, r.K, r.fpe, r.el, r.funding_cost,
            r.gross_margin, r.fpe_return, r.net_margin, r.raroc, r.raroc * 10000.0,
            r.t_end_years, r.df, r.revenue_pv, r.net_margin_pv, r.drawn_pv,
        ]
        for col, v in enumerate(vals, start=1):
            ws2.cell(row=ridx, column=col, value=v)

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 14
    ws2.column_dimensions["A"].width = 8

    # ── Sheet 3: Aggregates ─────────────────────────────────────
    ws3 = wb.create_sheet("Aggregates")
    agg_rows = [
        ("NPV borrower cost", aggregates["npv_borrower_cost"], "Σ revenue_i × DF_i"),
        ("NPV bank net margin", aggregates["npv_bank_net_margin"], "Σ net_margin_i × DF_i"),
        ("NPV drawn balance", aggregates["npv_drawn_balance"], "Σ avg_drawn_i × dt_i × DF_i"),
        ("Effective spread (decimal)", aggregates["effective_spread"],
         "NPV_borrower_cost / NPV_drawn_balance"),
        ("Effective spread (bp)", aggregates["effective_spread_bp"], "× 10,000"),
        ("FPE-weighted RAROC", aggregates["fpe_weighted_raroc"],
         "Σ raroc_i × FPE_i × dt_i / Σ FPE_i × dt_i"),
        ("Total revenue (undisc)", aggregates["total_revenue_undisc"], "Σ revenue_i"),
        ("Total EL (undisc)", aggregates["total_el_undisc"], "Σ EL_i"),
        ("Avg exposure (dt-weighted)", aggregates["avg_exposure"],
         "Σ exposure_i × dt_i / Σ dt_i"),
    ]
    ws3.cell(row=1, column=1, value="Metric")
    ws3.cell(row=1, column=2, value="Value")
    ws3.cell(row=1, column=3, value="Formula")
    for c in range(1, 4):
        _bold(ws3.cell(row=1, column=c))
    for ridx, (k, v, formula) in enumerate(agg_rows, start=2):
        ws3.cell(row=ridx, column=1, value=k)
        ws3.cell(row=ridx, column=2, value=v)
        ws3.cell(row=ridx, column=3, value=formula)
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 60

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)


# ──────────────────────────────────────────────────────────────────
# Driver
# ──────────────────────────────────────────────────────────────────

def build_one(deal, periods, cfg, discount, fixtures_dir, excel_dir):
    rows: List[PeriodRow] = []
    cum_t = 0.0
    for p in periods:
        r = compute_period(p, deal, cfg, cum_t, discount.rate)
        rows.append(r)
        cum_t = r.t_end_years
    aggs = compute_aggregates(rows)
    yaml_path = os.path.join(fixtures_dir, f"{deal.fixture_id}.yaml")
    excel_path = os.path.join(excel_dir, f"{deal.fixture_id}.xlsx")
    write_yaml(deal, periods, rows, aggs, cfg, discount, yaml_path)
    write_excel(deal, periods, rows, aggs, cfg, discount, excel_path)
    return rows, aggs, yaml_path, excel_path


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    excel_dir = os.path.join(here, "reference_excel")
    cfg = EngineConfigSnapshot()
    discount = DiscountSpec(kind="scalar", rate=cfg.risk_free_rate, day_count="Act/365F")

    print(f"engine_config = {asdict(cfg)}")
    print(f"discount = {asdict(discount)}")
    print()

    for builder in (fixture_rcf_5y, fixture_termloan_7y_amortising, fixture_projfin_10y_grace):
        deal, periods = builder()
        rows, aggs, yp, xp = build_one(deal, periods, cfg, discount, here, excel_dir)
        print(f"── {deal.fixture_id}")
        print(f"   {yp}")
        print(f"   {xp}")
        print(f"   {len(rows)} periods")
        for r in rows:
            print(f"   p{r.index}: rev={r.revenue:>12,.0f}  exp={r.exposure:>13,.0f}  "
                  f"K={r.K:.5f}  FPE={r.fpe:>12,.0f}  RAROC={r.raroc*100:>6.2f}%  DF={r.df:.5f}")
        print(f"   NPV borrower cost  : {aggs['npv_borrower_cost']:>14,.0f}")
        print(f"   NPV bank net margin: {aggs['npv_bank_net_margin']:>14,.0f}")
        print(f"   Effective spread   : {aggs['effective_spread_bp']:>10.2f} bp")
        print(f"   Weighted RAROC     : {aggs['fpe_weighted_raroc']*100:>10.2f} %")
        print()


if __name__ == "__main__":
    main()
